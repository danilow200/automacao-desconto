#------------------------------------------------------------------------------------------------------------------------
                                            #IMPORTAÇÕES
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
#------------------------------------------------------------------------------------------------------------------------
                                            #FUNÇÃO QUE FARÁ A ESTILIZAÇÃO DA SHEETS
#definie a função e o parâmetro
def estilizar_excel (nome_arquivo):
    df = pd.read_excel(nome_arquivo, sheet_name=None)
# Cria um objeto Workbook do openpyxl
    workbook = Workbook()
    # Remove a primeira planilha, que é criada automaticamente
    sheet_to_remove = workbook['Sheet']
    workbook.remove(sheet_to_remove)
# Itera sobre as sheets e estiliza cada uma delas
    for sheet_name, df in df.items():
        worksheet = workbook.create_sheet(sheet_name)
        # congelar as TRÊS primeiras linhas e todas as colunas
        worksheet.freeze_panes = 'A4'
        # for sheet in workbook.worksheets:
        worksheet.sheet_view.showGridLines = False
        # Cria um objeto Image com o caminho da imagem
        img = Image('.\logo.jpeg')
        # Definir o tamanho da imagem em centímetros
        img.width = 180.708
        img.height = 37
        # Adicionar a imagem à planilha
        worksheet.add_image(img, 'A1')
        # Define o estilo da fonte para o cabeçalho
        font = Font(bold=True, color='FFFFFF')

        # Define o estilo de alinhamento para o cabeçalho
        align = Alignment(horizontal='center', vertical='center')

        # Define o estilo de borda para as células
        #cabeçalho
        border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
        #resto
        border2 = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Define o estilo de preenchimento para as células
        fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
        fill_par = PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
        fill_impar = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')

        # Adiciona o cabeçalho na primeira linha da planilha
        for col_num, col_name in enumerate(df.columns, 2):
            cell = worksheet.cell(row=3, column=col_num, value=col_name)
            cell.font = font
            cell.alignment = align
            cell.border = border
            cell.fill = fill

        # Adiciona os dados na planilha
        for row_num, row_data in enumerate(df.values, 4):
            for col_num, cell_value in enumerate(row_data, 2):
                cell = worksheet.cell(row=row_num, column=col_num, value=cell_value)
                cell.alignment = align
                cell.border = border2
                #verifica se a celular é par ou impar para colocar determinada cor
                if row_num % 2 == 0:
                    cell.fill = fill_par
                else:
                    cell.fill = fill_impar

        # Itera sobre as colunas e ajusta a largura automaticamente
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # pega a letra da coluna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
#------------------------------------------------------------------------------------------------------------------------
                                            #FILTRO AUTOMÁTICO
        # Seleciona a linha que terá o filtro automático
        row_number = 3

        # Define o intervalo de células para serem filtradas
        cell_range = "B{}:{}{}".format(row_number, chr(ord('B')+df.shape[1]-1), row_number)

        # Adiciona o filtro automático para o intervalo de células selecionado
        worksheet.auto_filter.ref = cell_range
#------------------------------------------------------------------------------------------------------------------------        
   
    # Salva a planilha em um arquivo Excel
    workbook.save(nome_arquivo)
#------------------------------------------------------------------------------------------------------------------------