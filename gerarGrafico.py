import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_DATE_TIME6

def gerar_grafico(nome_arquivo):

    pagina = 'Comparação Total'
    
    workbook = openpyxl.load_workbook(nome_arquivo)

    worksheet = workbook[pagina]

    for row in worksheet.iter_rows(min_col=7, min_row=4, max_row=11):
        for cell in row:
            cell.number_format = FORMAT_PERCENTAGE_00 #converte float para porcetagem

    for row in worksheet.iter_rows(min_col=8, min_row=4, max_row=11):
        for cell in row:
            cell.number_format = FORMAT_PERCENTAGE_00

    for row in worksheet.iter_rows(min_col=3,max_col=6, min_row=4, max_row=11):
        for cell in row:
            cell.number_format = FORMAT_DATE_TIME6 #converte string para HH:MM:SS

    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Porcetagem PADTEC'
    chart.x_axis.title = 'Categorias'
    chart.y_axis.title = 'Porcetagem'

    categorias = Reference(worksheet, min_col=2, min_row=4, max_row=11)
    valores = Reference(worksheet, min_col=7, min_row=3, max_row=11)
    chart.add_data(valores, titles_from_data=True)
    chart.set_categories(categorias)
    chart.legend = None
    chart.show_legend = False
    chart.dataLabels = DataLabelList(showVal=True)

    chart_2 = BarChart()
    chart_2.type = 'col'
    chart_2.title = 'Porcetagem RADIANTE'
    chart_2.x_axis.title = 'Categorias'
    chart_2.y_axis.title = 'Porcetagem'

    categorias = Reference(worksheet, min_col=2, min_row=4, max_row=11)
    valores = Reference(worksheet, min_col=8, min_row=3, max_row=11)
    chart_2.add_data(valores, titles_from_data=True)
    chart_2.set_categories(categorias)
    chart_2.legend = None
    chart_2.show_legend = False
    chart_2.dataLabels = DataLabelList(showVal=True)

    worksheet.add_chart(chart, 'B14')
    worksheet.add_chart(chart_2, 'E14')

    workbook.save(nome_arquivo)